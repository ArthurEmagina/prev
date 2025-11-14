import datetime as dt
from dateutil import parser as dateparser
import pandas as pd
import streamlit as st
import yaml
import tempfile
import io

from pathlib import Path
from typing import Dict, Set, List, Tuple

import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Local imports
from calculator.logic import compute_offer
from calculator.data_loader import load_dataset

st.set_page_config(page_title="Calcul délais et coût", page_icon="📦", layout="wide")

@st.cache_data(show_spinner=False)
def read_config() -> dict:
	"""Charge la configuration depuis config.yaml"""
	config_path = Path(__file__).parent / "config.yaml"
	with open(config_path, "r", encoding="utf-8") as f:
		return yaml.safe_load(f)


@st.cache_data(show_spinner=True)
def load_dataset_cached(excel_path: str, config: dict):
	"""
	Charge le dataset avec cache basé sur le chemin.
	Le cache sera invalidé si un nouveau fichier est uploadé.
	"""
	# Mettre à jour le chemin dans la config
	config_copy = config.copy()
	config_copy["excel"] = config["excel"].copy()
	config_copy["excel"]["workbook_path"] = excel_path
	
	return load_dataset(config_copy)


def create_excel_template(config: dict, df: pd.DataFrame | None = None, template_type: str = "composants") -> bytes:
	"""
	Crée un fichier Excel template avec les colonnes exactes du fichier Excel.
	
	Args:
		config: Configuration du projet
		df: DataFrame du fichier Excel chargé (optionnel, pour récupérer les colonnes exactes)
		template_type: "composants" ou "fournisseurs" - détermine la première colonne
	"""
	wb = Workbook()
	ws = wb.active
	ws.title = config["excel"]["sheet_name"]
	
	# Si un DataFrame est fourni, utiliser ses colonnes exactes
	if df is not None and not df.empty:
		headers = list(df.columns)
	else:
		# Sinon, construire les colonnes à partir de la config
		columns_cfg = config["excel"]["columns"]
		product_col = columns_cfg.get("product_code", "Composants")
		moq_col = columns_cfg.get("moq", "MOQ")
		lot_col = columns_cfg.get("lot_size", "Lot")
		currency_col = columns_cfg.get("currency", "Devise")
		
		headers = [
			product_col,
			moq_col,
			lot_col,
		]
		
		# Colonnes optionnelles par palier (exemples de paliers courants)
		paliers = [200, 1000, 5000]
		
		# Ajouter les colonnes de coût unitaire de production par palier
		for palier in paliers:
			headers.append(f"CU pour {palier}")
		
		# Ajouter les colonnes de coût transport avion par palier
		for palier in paliers:
			headers.append(f"CTA pour {palier}")
		
		# Ajouter les colonnes de coût transport bateau par palier
		for palier in paliers:
			headers.append(f"CTB pour {palier}")
		
		# Ajouter les colonnes de temps de production par palier
		for palier in paliers:
			headers.append(f"Temps de prod pour {palier}")
		
		# Colonnes de délais de transport
		tiers_cfg = config["excel"].get("tiers", {})
		headers.append(tiers_cfg.get("air_transport_time_column", "Délais de transport Avion"))
		headers.append(tiers_cfg.get("sea_transport_time_column", "Délais de transport Bateau"))
		
		# Ajouter la colonne devise
		headers.append(currency_col)
	
	# Modifier la première colonne selon le type de template
	if template_type == "fournisseurs":
		headers[0] = "Fournisseurs"
	else:  # template_type == "composants"
		# Garder le nom original de la colonne (Composants ou autre)
		pass
	
	# Écrire les en-têtes
	for col_idx, header in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=col_idx, value=header)
		cell.font = Font(bold=True, color="FFFFFF")
		cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
		cell.alignment = Alignment(horizontal="center", vertical="center")
	
	# Ajouter une ligne d'exemple
	example_row = []
	for i, header in enumerate(headers):
		if i == 0:
			# Première colonne : Fournisseurs ou Composants
			if template_type == "fournisseurs":
				example_row.append("Exemple Fournisseur 1")
			else:
				example_row.append("Exemple Composant 1")
		elif "MOQ" in str(header) or "moq" in str(header).lower():
			example_row.append(100)
		elif "Lot" in str(header) or "lot" in str(header).lower():
			example_row.append(100)
		elif "CU pour" in str(header):
			# Extraire le palier pour donner un exemple cohérent
			example_row.append(5.50 if "200" in str(header) else (5.00 if "1000" in str(header) else 4.50))
		elif "CTA pour" in str(header):
			example_row.append(0.50 if "200" in str(header) else (0.50 if "1000" in str(header) else 0.45))
		elif "CTB pour" in str(header):
			example_row.append(0.20 if "200" in str(header) else (0.20 if "1000" in str(header) else 0.18))
		elif "Temps de prod" in str(header):
			example_row.append(2 if "200" in str(header) else (2 if "1000" in str(header) else 1.5))
		elif "Avion" in str(header) and "transport" in str(header):
			example_row.append(1)
		elif "Bateau" in str(header) and "transport" in str(header):
			example_row.append(4)
		elif "Devise" in str(header) or "currency" in str(header).lower():
			example_row.append("EUR")
		else:
			example_row.append("")  # Valeur vide pour les colonnes non reconnues
	
	for col_idx, value in enumerate(example_row, start=1):
		ws.cell(row=2, column=col_idx, value=value)
	
	# Ajuster la largeur des colonnes
	for col in ws.columns:
		max_length = 0
		col_letter = col[0].column_letter
		for cell in col:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		adjusted_width = min(max_length + 2, 30)
		ws.column_dimensions[col_letter].width = adjusted_width
	
	# Sauvegarder dans un buffer en mémoire
	buffer = io.BytesIO()
	wb.save(buffer)
	buffer.seek(0)
	return buffer.getvalue()


def _extract_assembly_days(df: pd.DataFrame, product_col: str, cfg: dict) -> int:
	asm_cfg = cfg.get("assembly", {}) or {}
	name = asm_cfg.get("component_name", "Assemblage et intégration final")
	col = asm_cfg.get("duration_weeks_column")
	fallback_weeks = int(asm_cfg.get("duration_weeks", 8) or 8)
	if col and col in df.columns:
		row = df[df[product_col].astype(str) == str(name)].head(1)
		if not row.empty:
			val = row.iloc[0][col]
			try:
				weeks = int(float(str(val).replace(",", ".")))
			except Exception:  # noqa: BLE001
				weeks = fallback_weeks
			return max(0, weeks) * 7
	return fallback_weeks * 7


def _resolve_dependencies(components: List[str], deps_map: Dict[str, List[str]]) -> Dict[str, List[str]]:
	"""Returns a dict: component -> list of all dependencies (recursive)"""
	all_deps: Dict[str, List[str]] = {}
	
	def get_all_deps(comp: str, visited: Set[str]) -> List[str]:
		if comp in visited:
			return []
		visited.add(comp)
		deps = deps_map.get(comp, []) or []
		result = []
		for dep in deps:
			result.append(dep)
			result.extend(get_all_deps(dep, visited))
		return result
	
	for comp in components:
		all_deps[comp] = get_all_deps(comp, set())
	return all_deps


def _backward_schedule_with_deps(components: List[str], deps_map: Dict[str, List[str]], assembly_start: dt.date, durations: Dict[str, int], assembly_end: dt.date | None = None, assembly_name: str | None = None) -> List[Dict]:
	# Build dependents map limited to selected components
	selected = set(components)
	dependents: Dict[str, List[str]] = {c: [] for c in components}
	for comp, deps in deps_map.items():
		if comp not in selected:
			continue
		for d in deps or []:
			if d in selected:
				dependents.setdefault(d, []).append(comp)
	# Initialize finish times: for nodes with no dependents, finish = assembly_start (or assembly_end for assembly)
	finish: Dict[str, dt.date] = {}
	order: List[str] = []
	# Topological order naive using DFS on dependents graph reversed (from sinks up)
	visited: Set[str] = set()
	def dfs(c: str):
		if c in visited:
			return
		visited.add(c)
		for nxt in dependents.get(c, []) or []:
			dfs(nxt)
		order.append(c)
	for c in components:
		dfs(c)
	# Compute backwards
	for c in order:
		deps_of_c = deps_map.get(c, []) or []
		deps_of_c = [d for d in deps_of_c if d in selected]
		dependents_of_c = dependents.get(c, []) or []
		if not dependents_of_c:
			# No dependents: this is a sink node
			# If it's assembly and we have assembly_end, use that; otherwise use assembly_start
			if assembly_name and c == assembly_name and assembly_end:
				finish[c] = assembly_end
			else:
				finish[c] = assembly_start
		else:
			# finish must be not later than min(start of all dependents)
			starts_of_dependents = []
			for depd in dependents_of_c:
				dur = int(durations.get(depd, 0))
				start_depd = finish.get(depd)
				if start_depd is None:
					continue
				start_depd = start_depd - dt.timedelta(days=dur)
				starts_of_dependents.append(start_depd)
			finish[c] = min(starts_of_dependents) if starts_of_dependents else assembly_start
	# Build rows: start = finish - duration
	rows: List[Dict] = []
	for c in components:
		dur = int(durations.get(c, 0))
		fin = finish.get(c, assembly_start)
		start = fin - dt.timedelta(days=dur)
		rows.append({"Composant": c, "Start": start, "Finish": fin})
	return rows


def _forward_schedule_with_custom_starts(
	components: List[str],
	deps_map: Dict[str, List[str]],
	start_by_comp: Dict[str, dt.date],
	lead_days: Dict[str, int],
	assembly_name: str,
	assembly_days: int,
) -> List[Dict]:
	"""Forward schedule with per-component start dates (order dates) and dependencies.
	Each task starts at max(own start, finishes of its dependencies). Returns rows dicts.
	"""
	selected = set(components)
	deps_limited: Dict[str, List[str]] = {c: [d for d in (deps_map.get(c, []) or []) if d in selected] for c in components}
	starts: Dict[str, dt.date] = {}
	finishes: Dict[str, dt.date] = {}
	remaining = set(components)
	iteration = 0
	max_iter = len(components) * 3
	while remaining and iteration < max_iter:
		iteration += 1
		progress = False
		for comp in list(remaining):
			deps = deps_limited.get(comp, [])
			if all(d in finishes for d in deps):
				own_start = start_by_comp.get(comp) or dt.date.today()
				if deps:
					latest_dep_fin = max(finishes[d] for d in deps)
					start = max(own_start, latest_dep_fin)
				else:
					start = own_start
				lead = int(lead_days.get(comp, 0))
				finish = start + dt.timedelta(days=max(0, lead))
				starts[comp] = start
				finishes[comp] = finish
				remaining.remove(comp)
				progress = True
		if not progress:
			for comp in list(remaining):
				own_start = start_by_comp.get(comp) or dt.date.today()
				deps = deps_limited.get(comp, [])
				if deps:
					dep_finishes = [finishes.get(d) for d in deps if d in finishes]
					if dep_finishes:
						latest_dep_fin = max(dep_finishes)
						start = max(own_start, latest_dep_fin)
					else:
						start = own_start
				else:
					start = own_start
				lead = int(lead_days.get(comp, 0))
				starts[comp] = start
				finishes[comp] = start + dt.timedelta(days=max(0, lead))
				remaining.remove(comp)
			break
	rows: List[Dict] = []
	for comp in components:
		if comp in starts and comp in finishes:
			rows.append({"Composant": comp, "Start": starts[comp], "Finish": finishes[comp]})
	# Add production phase (assembly) after all deps
	asm_deps = [d for d in (deps_map.get(assembly_name, []) or []) if d in selected]
	if (assembly_name in selected) or (asm_deps and all(d in finishes for d in asm_deps)):
		asm_start = max((finishes[d] for d in asm_deps), default=max(finishes.values()) if finishes else dt.date.today())
		asm_finish = asm_start + dt.timedelta(days=assembly_days)
		rows.append({"Composant": assembly_name, "Start": asm_start, "Finish": asm_finish})
	return rows

def main() -> None:
	st.title("Calcul des délais et coûts – Avion vs Bateau")
	st.markdown("🌐 **Version en ligne** - Accessible depuis n'importe quel appareil")

	# Sidebar: configuration
	st.sidebar.header("Configuration")
	
	# Charger la config
	try:
		config = read_config()
	except Exception as exc:  # noqa: BLE001
		st.error(f"Impossible de lire la configuration: {exc}")
		st.stop()

	# Sélecteur de fichier Excel (uniquement upload)
	st.sidebar.subheader("Fichier Excel")
	
	# Option pour créer des fichiers (disponible même sans fichier chargé)
	with st.sidebar.expander("📝 Création de fichiers", expanded=False):
		st.caption("Génère un fichier Excel prérempli avec toutes les colonnes nécessaires")
		st.info(
			"💡 Si vous n'avez pas encore créé de fichier Excel, "
			"téléchargez d'abord un des fichiers préremplis ci-dessous."
		)
		
		template_type = st.radio(
			"Type de fichier",
			options=["composants", "fournisseurs"],
			format_func=lambda x: "Composants" if x == "composants" else "Fournisseurs",
			key="template_type_choice"
		)
		
		if st.button("🔄 Générer le fichier", use_container_width=True, key="generate_template"):
			try:
				# Utiliser None pour df si pas encore chargé (utilisera la config)
				template_bytes = create_excel_template(config, None, template_type)
				st.session_state[f'template_bytes_{template_type}'] = template_bytes
				st.success("✓ Fichier créé !")
			except Exception as e:
				st.error(f"Erreur: {e}")
		
		# Boutons de téléchargement pour chaque type
		if f'template_bytes_composants' in st.session_state:
			st.download_button(
				label="⬇️ Télécharger Fichier Composants.xlsx",
				data=st.session_state['template_bytes_composants'],
				file_name="InputDélais_Composants.xlsx",
				mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
				use_container_width=True,
				key="download_composants"
			)
		
		if f'template_bytes_fournisseurs' in st.session_state:
			st.download_button(
				label="⬇️ Télécharger Fichier Fournisseurs.xlsx",
				data=st.session_state['template_bytes_fournisseurs'],
				file_name="InputDélais_Fournisseurs.xlsx",
				mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
				use_container_width=True,
				key="download_fournisseurs"
			)
	
	uploaded_file = st.sidebar.file_uploader(
		"📤 Téléchargez votre fichier Excel",
		type=['xlsm', 'xlsx'],
		help="Sélectionnez InputDélais.xlsm depuis votre ordinateur"
	)
	
	# Vérifier qu'un fichier a été uploadé
	if uploaded_file is None:
		st.info("👆 Veuillez télécharger votre fichier Excel dans la barre latérale pour commencer.")
		st.markdown("""
		### 📋 Instructions
		1. Si vous n'avez pas encore de fichier Excel, téléchargez d'abord un fichier prérempli dans la section **"Création de fichiers"** ci-dessus
		2. Remplissez votre fichier Excel avec vos données
		3. Cliquez sur **"Browse files"** dans la barre latérale
		4. Sélectionnez votre fichier `InputDélais.xlsm`
		5. L'application chargera automatiquement les données
		""")
		st.stop()
	
	# Sauvegarder le fichier uploadé temporairement
	temp_dir = Path(tempfile.gettempdir()) / "calcul_delais_cout_cloud"
	try:
		temp_dir.mkdir(exist_ok=True)
	except (OSError, PermissionError) as e:
		st.error(f"Impossible de créer le dossier temporaire: {e}")
		st.stop()
	
	temp_path = temp_dir / uploaded_file.name
	
	# Vérifier si c'est un nouveau fichier (pour invalider le cache)
	file_id = f"{uploaded_file.name}_{uploaded_file.size}"
	if 'last_file_id' not in st.session_state or st.session_state.last_file_id != file_id:
		# Nouveau fichier, invalider le cache
		load_dataset_cached.clear()
		st.session_state.last_file_id = file_id
	
	# Sauvegarder le fichier
	try:
		with open(temp_path, "wb") as f:
			f.write(uploaded_file.getbuffer())
		excel_path = str(temp_path.resolve())
		st.sidebar.success(f"✓ Fichier chargé: {uploaded_file.name}")
	except (OSError, PermissionError) as e:
		st.error(f"Impossible d'écrire le fichier temporaire: {e}")
		st.stop()
	
	# Bouton de rafraîchissement manuel
	if st.sidebar.button("🔄 Recharger les données", use_container_width=True):
		load_dataset_cached.clear()
		st.rerun()
	
	# Load dataset avec cache
	try:
		excel_path_obj = Path(excel_path)
		if not excel_path_obj.is_file():
			st.error(f"⚠️ Le fichier n'est pas valide: {excel_path}")
			st.stop()
		
		df, mapping = load_dataset_cached(excel_path, config)
		st.sidebar.success("✓ Données chargées")
	except PermissionError as exc:
		st.error(f"⚠️ Erreur de permission: {exc}")
		st.info("Assurez-vous que le fichier Excel n'est pas corrompu.")
		st.stop()
	except FileNotFoundError as exc:
		st.error(f"⚠️ Fichier introuvable: {exc}")
		st.stop()
	except Exception as exc:  # noqa: BLE001
		st.error(f"Erreur lors du chargement du fichier Excel: {exc}")
		st.code(f"Type d'erreur: {type(exc).__name__}\nDétails: {str(exc)}", language="text")
		st.stop()

	# Vérifier que les données sont bien chargées
	if df is None or df.empty:
		st.error("⚠️ Le fichier Excel est vide ou n'a pas pu être chargé correctement.")
		st.stop()
	
	if not mapping or "product_code" not in mapping:
		st.error("⚠️ Configuration incorrecte: colonne 'product_code' introuvable.")
		st.stop()
	
	# Option pour créer des fichiers avec les colonnes exactes (après le chargement du fichier)
	# Cette section permet de régénérer avec les colonnes exactes du fichier chargé
	with st.sidebar.expander("📝 Régénérer avec colonnes exactes", expanded=False):
		st.caption("Génère un fichier avec les colonnes exactes de votre fichier chargé")
		
		template_type_exact = st.radio(
			"Type de fichier",
			options=["composants", "fournisseurs"],
			format_func=lambda x: "Composants" if x == "composants" else "Fournisseurs",
			key="template_type_choice_exact"
		)
		
		if st.button("🔄 Régénérer le fichier", use_container_width=True, key="regenerate_template"):
			try:
				template_bytes = create_excel_template(config, df, template_type_exact)
				st.session_state[f'template_bytes_{template_type_exact}'] = template_bytes
				st.success("✓ Fichier régénéré !")
			except Exception as e:
				st.error(f"Erreur: {e}")

	# Détection automatique du type de fichier (Composants ou Fournisseurs)
	first_col_name = df.columns[0] if not df.empty else ""
	product_col = mapping["product_code"]
	
	# Vérifier si la première colonne correspond à "Fournisseurs" ou "Composants"
	is_fournisseurs_mode = (
		"fournisseur" in first_col_name.lower() or 
		first_col_name.lower() == "fournisseurs" or
		first_col_name.lower() == "fournisseur"
	)
	
	# Si c'est le mode fournisseurs, utiliser la première colonne, sinon utiliser product_col
	if is_fournisseurs_mode:
		selection_col = first_col_name
		selection_label = "Fournisseurs"
		item_label = "fournisseur"
	else:
		selection_col = product_col
		selection_label = "Composants"
		item_label = "composant"
	
	# Récupérer la liste des éléments à sélectionner (composants ou fournisseurs)
	all_items = sorted(df[selection_col].astype(str).dropna().unique().tolist())
	
	# Pour les calculs, on utilise toujours product_col (qui peut être la même que selection_col)
	all_components = sorted(df[product_col].astype(str).dropna().unique().tolist()) if product_col in df.columns else all_items
	
	# Vérifier que la colonne existe dans le DataFrame
	if product_col not in df.columns:
		st.error(f"⚠️ La colonne '{product_col}' n'existe pas dans le fichier Excel.")
		st.info(f"Colonnes disponibles: {', '.join(df.columns.tolist()[:10])}...")
		st.stop()
	
	# Vérifier qu'il y a des éléments
	if not all_items:
		st.warning(f"⚠️ Aucun {item_label} trouvé dans le fichier Excel.")
		st.info(f"Vérifiez que la colonne des {selection_label.lower()} contient des données.")
		st.stop()
	
	# Debug: afficher le nombre d'éléments
	st.sidebar.caption(f"📊 {len(all_items)} {item_label}(s) trouvé(s)")
	
	assembly_days = _extract_assembly_days(df, product_col, config)
	assembly_weeks = assembly_days // 7
	deps_map: Dict[str, List[str]] = config.get("dependencies", {}) or {}
	assembly_name = (config.get("assembly", {}) or {}).get("component_name", "Assemblage et intégration final")

	# Tabs
	plan_tab, suivi_tab = st.tabs(["Plan par mode", "Suivi de commande"])

	with plan_tab:
		# Inputs généraux
		colp1, colp2 = st.columns(2)
		with colp1:
			quantity_p = st.number_input(f"Quantité par défaut (par {item_label})", min_value=1, step=1, value=100, key="qty_plan", help="Utilisée si aucune quantité spécifique n'est renseignée")
		with colp2:
			target_delivery_date = st.date_input("Date de livraison cible (fin)", value=dt.date.today() + dt.timedelta(days=30), key="target_delivery_plan")

		selected_p = st.multiselect(selection_label, options=all_items, help=f"Cochez un ou plusieurs {selection_label.lower()}", key="sel_plan")

		if not selected_p:
			st.info(f"Sélectionnez au moins un {item_label} pour planifier.")
		else:
			st.divider()
			st.subheader(f"Configuration par {item_label}")
			
			# For each selected component: name, quantity, mode
			choices: Dict[str, str] = {}
			lead_days_by_comp: Dict[str, int] = {}
			mode_by_comp: Dict[str, str] = {}
			cost_by_comp: Dict[str, float] = {}
			quantity_by_comp: Dict[str, int] = {}
			currency2 = None
			
			for comp in selected_p:
				with st.container():
					st.markdown(f"### {comp}")
					
					# Colonnes pour quantité et mode
					col_qty, col_mode = st.columns([1, 1])
					
					with col_qty:
						qty_input = st.number_input(
							"Quantité",
							min_value=0,
							step=1,
							value=quantity_p,
							key=f"qty_comp_{comp}",
							help=f"Modifiez si nécessaire (par défaut: {quantity_p})"
						)
						# Si 0, utiliser la quantité générale, sinon utiliser la valeur saisie
						qty_for_comp = qty_input if qty_input > 0 else quantity_p
						quantity_by_comp[comp] = qty_for_comp
					
					# Calculer les résultats pour ce composant/fournisseur
					# Utiliser selection_col pour trouver la ligne, mais product_col pour les calculs
					if is_fournisseurs_mode and selection_col != product_col:
						# En mode fournisseurs, chercher par la colonne de sélection
						row_df = df[df[selection_col].astype(str) == str(comp)].head(1)
					else:
						# En mode composants, utiliser product_col
						row_df = df[df[product_col].astype(str) == str(comp)].head(1)
					
					if not row_df.empty:
						res = compute_offer(row_df.iloc[0], qty_for_comp, target_delivery_date - dt.timedelta(days=assembly_days), mapping, config)
						currency2 = currency2 or res.get("currency", "EUR")
						lead_air = int(res['air']['lead_time_days'])
						lead_sea = int(res['sea']['lead_time_days'])
						cost_air = float(res['air']['total_cost'])
						cost_sea = float(res['sea']['total_cost'])
						
						with col_mode:
							if comp != assembly_name and lead_air != lead_sea:
								choice = st.radio(
									"Mode de transport",
									options=["Avion", "Bateau"],
									horizontal=True,
									key=f"mode_{comp}",
									help=f"Avion: {lead_air}j, {cost_air:.2f} {currency2 or 'EUR'} | Bateau: {lead_sea}j, {cost_sea:.2f} {currency2 or 'EUR'}"
								)
							else:
								# Pas de choix si délais identiques ou pour l'assemblage: on fige sur le moins cher (ou avion par défaut si égalité de coût)
								choice = "Avion" if cost_air <= cost_sea else "Bateau"
								st.radio(
									"Mode de transport",
									options=[choice],
									index=0,
									key=f"mode_{comp}",
									disabled=True,
									help=f"Mode fixé: {choice} (délais identiques ou assemblage)"
								)
								st.caption(f"Mode fixé: {choice} (délais identiques)")
							
							choices[comp] = choice
							if choice == "Avion":
								lead_days_by_comp[comp] = lead_air
								mode_by_comp[comp] = "Avion"
								cost_by_comp[comp] = cost_air
							else:
								lead_days_by_comp[comp] = lead_sea
								mode_by_comp[comp] = "Bateau"
								cost_by_comp[comp] = cost_sea
					
					st.divider()

			# Determine assembly end date: if assembly is selected, it ends at target; otherwise compute backward
			if assembly_name in selected_p:
				assembly_end_date = target_delivery_date
				assembly_start_date = target_delivery_date - dt.timedelta(days=assembly_days)
			else:
				assembly_end_date = target_delivery_date
				assembly_start_date = target_delivery_date - dt.timedelta(days=assembly_days)

			total_cost = sum(cost_by_comp.get(c, 0.0) for c in selected_p)

			# Backward schedule: components must finish by assembly_start_date, assembly ends at target_delivery_date
			rows_sched = _backward_schedule_with_deps(selected_p, deps_map, assembly_start_date, lead_days_by_comp, assembly_end_date if assembly_name in selected_p else None, assembly_name if assembly_name in selected_p else None)
			if not rows_sched:
				st.warning("Aucune planification générée.")
			else:
				# Find earliest start (when to start ordering)
				earliest_start = min(r['Start'] for r in rows_sched) if rows_sched else None
				overall_arrival = max(r['Finish'] for r in rows_sched) if rows_sched else None
				
				# Final delivery is always the target date (user input)
				final_delivery = target_delivery_date

				st.divider()
				mA, sA, mB, sB, mC = st.columns([1, 0.3, 1, 0.3, 1])
				with mA:
					st.metric("Coût total (sélection)", f"{total_cost:.2f} {currency2 or 'EUR'}")
				with mB:
					st.metric("Commander au plus tard", earliest_start.isoformat() if earliest_start else "-")
				with mC:
					st.metric("Date de livraison (fin)", final_delivery.isoformat() if final_delivery else "-")

				# Gantt chart with dependencies
				st.subheader("Roadmap (Gantt)")
				gantt_data = []
				for r in rows_sched:
					comp = r['Composant']
					lead = int(lead_days_by_comp.get(comp, 0))
					cost = float(cost_by_comp.get(comp, 0.0))
					start = r['Start']
					finish = r['Finish']
					# Ensure at least 1 day visibility
					if finish <= start:
						finish = start + dt.timedelta(days=max(1, lead))
					gantt_data.append({
						"Composant": comp,
						"Lead (j)": lead,
						"Coût": cost,
						"Start": start,
						"Finish": finish,
					})
				
				if not gantt_data:
					st.warning("Aucune donnée à afficher dans le Gantt.")
				else:
					gdf2 = pd.DataFrame(gantt_data)
					fig2 = px.timeline(gdf2, x_start="Start", x_end="Finish", y="Composant", hover_data=["Lead (j)", "Coût"])
					fig2.update_yaxes(autorange="reversed")
					fig2.update_layout(showlegend=False)
					st.plotly_chart(fig2, use_container_width=True)
					
					with st.expander("Détails planning (par composant)"):
						st.dataframe(gdf2[["Composant", "Lead (j)", "Coût", "Start", "Finish"]], use_container_width=True, hide_index=True)

	with suivi_tab:
		# Inputs généraux
		col_s1, col_s2 = st.columns(2)
		with col_s1:
			quantity_s = st.number_input(f"Quantité par défaut (par {item_label})", min_value=1, step=1, value=100, key="qty_suivi", help="Utilisée si aucune quantité spécifique n'est renseignée")
		with col_s2:
			today = dt.date.today()
			st.write(f"Aujourd'hui: {today.isoformat()}")

		selected_s = st.multiselect(selection_label, options=all_items, help=f"Cochez un ou plusieurs {selection_label.lower()}", key="sel_suivi")
		if not selected_s:
			st.info(f"Sélectionnez au moins un {item_label} pour le suivi.")
		else:
			st.divider()
			st.subheader(f"Configuration par {item_label}")
			
			mode_by_comp_s: Dict[str, str] = {}
			lead_days_s: Dict[str, int] = {}
			order_date_by_comp: Dict[str, dt.date] = {}
			quantity_by_comp_s: Dict[str, int] = {}
			currency3 = None
			
			for comp in selected_s:
				with st.container():
					st.markdown(f"### {comp}")
					
					# Colonnes pour quantité, mode et date de commande
					col_qty, col_mode, col_date = st.columns([1, 1, 1])
					
					with col_qty:
						qty_input = st.number_input(
							"Quantité",
							min_value=0,
							step=1,
							value=quantity_s,
							key=f"qty_suivi_{comp}",
							help=f"Modifiez si nécessaire (par défaut: {quantity_s})"
						)
						# Si 0, utiliser la quantité générale, sinon utiliser la valeur saisie
						qty_for_comp = qty_input if qty_input > 0 else quantity_s
						quantity_by_comp_s[comp] = qty_for_comp
					
					# Calculer les résultats pour ce composant/fournisseur
					# Utiliser selection_col pour trouver la ligne, mais product_col pour les calculs
					if is_fournisseurs_mode and selection_col != product_col:
						# En mode fournisseurs, chercher par la colonne de sélection
						row_df = df[df[selection_col].astype(str) == str(comp)].head(1)
					else:
						# En mode composants, utiliser product_col
						row_df = df[df[product_col].astype(str) == str(comp)].head(1)
					
					if not row_df.empty:
						res = compute_offer(row_df.iloc[0], qty_for_comp, today, mapping, config)
						currency3 = currency3 or res.get("currency", "EUR")
						lead_air = int(res['air']['lead_time_days'])
						lead_sea = int(res['sea']['lead_time_days'])
						cost_air = float(res['air']['total_cost'])
						cost_sea = float(res['sea']['total_cost'])
						
						with col_mode:
							if comp != assembly_name and lead_air != lead_sea:
								choice = st.radio(
									"Mode de transport",
									options=["Avion", "Bateau"],
									horizontal=True,
									key=f"mode_s_{comp}",
									help=f"Avion: {lead_air}j, {cost_air:.2f} {currency3 or 'EUR'} | Bateau: {lead_sea}j, {cost_sea:.2f} {currency3 or 'EUR'}"
								)
							else:
								# Pas de choix si délais identiques ou pour l'assemblage: on fige sur le moins cher (ou avion par défaut si égalité de coût)
								choice = "Avion" if cost_air <= cost_sea else "Bateau"
								st.radio(
									"Mode de transport",
									options=[choice],
									index=0,
									key=f"mode_s_{comp}",
									disabled=True,
									help=f"Mode fixé: {choice} (délais identiques ou assemblage)"
								)
								st.caption(f"Mode fixé: {choice} (délais identiques)")
							
							mode_by_comp_s[comp] = choice
							lead_days_s[comp] = lead_air if choice == "Avion" else lead_sea
						
						with col_date:
							order_date_by_comp[comp] = st.date_input(
								"Date de commande",
								value=today,
								key=f"order_{comp}"
							)
					
					st.divider()

			rows_fwd = _forward_schedule_with_custom_starts(selected_s, deps_map, order_date_by_comp, lead_days_s, assembly_name, assembly_days)
			if not rows_fwd:
				st.warning("Aucune planification générée.")
			else:
				gantt_rows = []
				for r in rows_fwd:
					comp = r["Composant"]
					lead = int(lead_days_s.get(comp, 0))
					start = r["Start"]
					finish = r["Finish"]
					if finish <= start:
						finish = start + dt.timedelta(days=max(1, lead))
					restant = max(0, (finish - today).days)
					gantt_rows.append({
						"Composant": comp,
						"Lead (j)": lead,
						"Restant (j)": restant,
						"Start": start,
						"Finish": finish,
					})
				gdf3 = pd.DataFrame(gantt_rows)
				st.subheader("Suivi – temps restant (Gantt)")
				fig3 = px.timeline(gdf3, x_start="Start", x_end="Finish", y="Composant", hover_data=["Lead (j)", "Restant (j)"])
				fig3.update_yaxes(autorange="reversed")
				# Today marker
				today_dt = dt.datetime.combine(today, dt.time())
				fig3.add_shape(type="line", x0=today_dt, x1=today_dt, y0=0, y1=1, xref="x", yref="paper", line=dict(color="black", dash="dot"))
				fig3.update_layout(showlegend=False)
				st.plotly_chart(fig3, use_container_width=True)
				with st.expander("Détails (par composant)"):
					st.dataframe(gdf3[["Composant", "Lead (j)", "Restant (j)", "Start", "Finish"]], use_container_width=True, hide_index=True)


if __name__ == "__main__":
	main()
